"""
Homework 8, Exercise 4
William Morris
3/17/19
This program converts excel files and spreadsheets into CSV files. Excel spreadsheets in the file
"excelSpreadsheets" are converted csv files where each sheet of the spreadsheet is it's on csv file. CSV files are
placed in the file "csvFiles". File locations are based on the current directory.
"""

import os
import openpyxl
import csv
import shutil

# create the csvFiles folder for the converted files
csvFiles = os.path.join(".", "csvFiles")
if os.path.exists(csvFiles):
    shutil.rmtree(csvFiles)
os.mkdir(csvFiles)

# create the path for the excel files in the excelSpreadsheets folder
excelFiles = os.path.join(".", "excelSpreadsheets")

# Loop through each excel file in the folder and convert each excel sheet to its own csv file
for excelFile in os.listdir(excelFiles):
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(os.path.join(excelFiles, excelFile))
    # Skip non-xlsx files, load the workbook object
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the excel filename and sheet title.
        csvFile = str(excelFile)[:-5] + '_' + sheet.title + ".csv"

        # Create the csv.writer object for this CSV file.
        csvFilename = open(os.path.join(csvFiles, csvFile), 'w', newline='')
        csvFileWriter = csv.writer(csvFilename)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            # append each cell to this list
            rowData = []
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            # Write the rowData list to the CSV file.
            csvFileWriter.writerow(rowData)
        csvFilename.close()
