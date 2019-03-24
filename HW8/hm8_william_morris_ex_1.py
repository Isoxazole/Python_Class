"""
Homework 8, Exercise 1
William Morris
3/17/19
This program inverts the column and row of the cells in a spreadsheet.
"""

import os
import openpyxl

# This will delete the inverted file if already made
ext = ["invertedExample.xlsx"]
[os.remove(filename) for filename in os.listdir(".") if filename in ext]

# Open workbook and set as active sheet
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb.get_active_sheet()

# Define max row and column
maxRow = sheet.max_row + 1
maxColumn = sheet.max_column + 1

# Append all values in the excel sheet to a 2d array and clear values
values = []
for i in range(1, maxRow):
    rowValues = []
    for j in range(1, maxColumn):
        rowValues.append(sheet.cell(row=i, column=j).value)
        sheet.cell(row=i, column=j).value = None
    values.append(rowValues)
# Check values are correctly in 2d array
print(values)

# Invert the values in the 2d array when putting back in sheet
# (cell sizes remain the same size)
for i in range(1,maxRow):
    for j in range(1, maxColumn):
        sheet.cell(row=j, column=i).value = values[i-1][j-1]

# Save the file
wb.save("invertedExample.xlsx")
wb.close()
